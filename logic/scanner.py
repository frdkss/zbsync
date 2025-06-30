from pathlib import Path
from typing import List, Optional
from openpyxl import load_workbook

from loging.cfg import logger
from .cfg import ZUP_KEYS, BUH_KEYS, ZUP_TEXT, BUH_TEXT, CODES
from .models import FileMeta
from .utils import parse_date, load_table


def scan(root: str) -> List[FileMeta]:
    logger.info("scan: enter root='{}'", root)
    if not Path(root).is_dir():
        logger.critical("scan: root '{}' не является директорией", root)
        return []

    metas: List[FileMeta] = []
    for f in Path(root).rglob("*.xlsx"):
        logger.info("scan: found file '{}'", f)
        if f.name.startswith("~$"):
            logger.info("scan: skip temp file '{}'", f)
            continue

        try:
            wb = load_workbook(str(f), read_only=True, data_only=True)
            logger.info("scan: opened '{}'", f)
        except Exception:
            logger.exception("scan: cannot open '{}'", f)
            logger.warning("scan: пропускаем файл '{}'", f)
            continue

        try:
            ws0 = wb.active
            head_cells = [ws0["D4"].value, ws0["E4"].value]
            all_texts = [
                cell
                for ws in wb.worksheets
                for row in ws.iter_rows(values_only=True)
                for cell in row if isinstance(cell, str)
            ]
            logger.info("scan: collected {} text items", len(all_texts))
        finally:
            wb.close()
            logger.info("scan: closed '{}'", f)

        is_zup = any(k in f.name.lower() for k in ZUP_KEYS) \
                 or any(ZUP_TEXT.lower() in t.lower() for t in all_texts)
        is_buh = any(k in f.name.lower() for k in BUH_KEYS) \
                 or any(BUH_TEXT.lower() in t.lower() for t in all_texts)
        logger.info("scan: tags for '{}' → ZUP={}, BUH={}", f, is_zup, is_buh)
        if not (is_zup or is_buh):
            logger.warning("scan: '{}' не содержит ЗУП/БУХ → skip", f)
            continue

        logger.info("scan: calling _find_date for '{}'", f)
        dt = _find_date(str(f), head_cells, all_texts)
        logger.info("scan: _find_date returned {}", dt)
        if dt is None:
            logger.warning("scan: в '{}' не найдена дата → skip", f)
            continue

        m, d = dt.month, dt.day
        q = (m - 1) // 3 + 1
        idx = (m - 1) % 3 + 1
        term = 1 if d < 22 else 2
        code = CODES[idx][term]
        logger.info("scan: computed m={}, d={}, q={}, idx={}, term={}, code={}", m, d, q, idx, term, code)

        table = load_table(str(f))
        if not table:
            logger.warning("scan: таблица из '{}' пуста", f)

        metas.append(FileMeta(
            path=str(f),
            tags=(["ЗУП"] if is_zup else []) + (["БУХ"] if is_buh else []),
            date=dt, month=m, day=d,
            quarter=q, idx=idx, term=term,
            code=code, table=table
        ))
        logger.info("scan: appended FileMeta for '{}'", f)

    logger.info("scan: exit with {} items", len(metas))
    return metas


def _find_date(path: str, heads: List[Optional[object]], texts: List[str]) -> Optional["date"]:
    from datetime import date
    logger.info("_find_date: enter path='{}'", path)
    wb = None
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        for loc in ("D4", "E4"):
            dt = parse_date(ws[loc].value)
            logger.info("_find_date: parsed {} at {} -> {}", path, loc, dt)
            if dt:
                logger.info("_find_date: exit with {} from {}", dt, loc)
                return dt
    except Exception:
        logger.warning("_find_date: не удалось прочитать D4/E4 в '{}'", path)
    finally:
        if wb:
            wb.close()
            logger.info("_find_date: closed '{}'", path)

    for t in texts:
        dt = parse_date(t)
        logger.info("_find_date: trying text {!r} -> {}", t, dt)
        if dt:
            logger.info("_find_date: exit with {} from text", dt)
            return dt

    logger.warning("_find_date: не удалось найти дату в текстах для '{}'", path)
    return None
