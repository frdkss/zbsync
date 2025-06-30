import re
from datetime import date
from typing import Any, List, Optional
from openpyxl import load_workbook
from loging.cfg import logger


def parse_date(value: Any) -> Optional[date]:
    logger.info("parse_date: enter value={!r}", value)
    try:
        if hasattr(value, "date"):
            result = value.date()
            logger.info("parse_date: .date() -> {}", result)
            return result
        if isinstance(value, str):
            m = re.search(r"(\d{1,2})\.(\d{1,2})\.(\d{2,4})", value)
            if m:
                d, mth, yr = map(int, m.groups())
                if yr < 100:
                    yr += 2000
                result = date(yr, mth, d)
                logger.info("parse_date: parsed string -> {}", result)
                return result
    except Exception:
        logger.exception("parse_date: failed to parse {!r}", value)

    logger.warning("parse_date: не удалось извлечь дату из {!r}", value)
    return None


def load_table(path: str, header: str = "Сотрудник, СНИЛС") -> List[List[Any]]:
    logger.info("load_table: enter path='{}', header='{}'", path, header)
    wb = None
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        logger.info("load_table: workbook opened '{}'", path)
    except Exception:
        logger.exception("load_table: cannot open workbook '{}'", path)
        logger.warning("load_table: возвращаю пустую таблицу для '{}'", path)
        return []

    try:
        ws = wb.active
        header_row: Optional[int] = next(
            (i for i, row in enumerate(ws.iter_rows(values_only=True), 1)
             if any(isinstance(c, str) and c.strip() == header for c in row)),
            None
        )
        if header_row is None:
            logger.warning("load_table: заголовок '{}' не найден в '{}'", header, path)
            return []

        logger.info("load_table: header found at row {}", header_row)
        rows2 = list(ws.iter_rows(min_row=header_row, max_row=header_row + 1, values_only=True))
        max_col = max(
            max((j for j, v in enumerate(rows2[0]) if v is not None), default=-1),
            max((j for j, v in enumerate(rows2[1]) if v is not None), default=-1)
        ) + 1
        logger.info("load_table: max_col = {}", max_col)

        last = header_row
        for r in range(ws.max_row, header_row - 1, -1):
            if any(ws.cell(r, c + 1).value is not None for c in range(max_col)):
                last = r
                break
        logger.info("load_table: last data row = {}", last)

        table = [
            [ws.cell(r, c + 1).value for c in range(max_col)]
            for r in range(header_row, last + 1)
        ]
        logger.info("load_table: extracted {} rows", len(table))
        return table

    except Exception:
        logger.exception("load_table: error reading table from '{}'", path)
        logger.warning("load_table: возвращаю пустую таблицу для '{}'", path)
        return []
    finally:
        if wb:
            wb.close()
            logger.info("load_table: workbook closed '{}'", path)
