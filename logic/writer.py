from datetime import datetime, date
from typing import List
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.formatting.rule import CellIsRule

from loging.cfg import logger
from .cfg import TEMPLATE, OUTPUT, MONTHS
from .models import FileMeta


def export_report(metas: List[FileMeta]) -> str:
    logger.info("export_report: enter with {} items", len(metas))

    groups: dict[str, dict] = {}
    for m in metas:
        title = f"{MONTHS[m.month]}_{m.code}_{m.quarter}квартал"
        grp = groups.setdefault(title, {
            "month": m.month, "term": m.term, "code": m.code, "recs": []
        })
        for tag in m.tags:
            grp["recs"].append((tag, m.table))
    logger.info("export_report: grouped into {} sheets", len(groups))

    wb = None
    try:
        wb = load_workbook(TEMPLATE)
        logger.info("export_report: opened template '{}'", TEMPLATE)
    except Exception:
        logger.critical("export_report: не удалось открыть шаблон '{}', прерываю", TEMPLATE)
        raise

    base = wb.sheetnames[0]
    red = Font(color="FFFF0000")

    for title, grp in groups.items():
        logger.info("export_report: start sheet '{}'", title)
        try:
            ws = wb.copy_worksheet(wb[base])
            ws.title = title

            sd = 1 if grp["term"] == 1 else 22
            ws["A1"].value = date(datetime.now().year, grp["month"], sd).strftime("%d.%m.%Y")
            ws["E1"].value = ws["O1"].value = grp["code"]
            ws["A2"].value = "БУХ";
            ws["K2"].value = "ЗУП"
            logger.info("export_report: header done for '{}'", title)

            row0 = 6
            buh = [t for tag, t in grp["recs"] if tag == "БУХ"]
            zup = [t for tag, t in grp["recs"] if tag == "ЗУП"]
            tb_b = buh[0] if buh else []
            tb_z = zup[0] if zup else []

            if not tb_b:
                logger.warning("export_report: нет данных БУХ для листа '{}'", title)
            if not tb_z:
                logger.warning("export_report: нет данных ЗУП для листа '{}'", title)

            for i, row in enumerate(tb_b):
                for j, v in enumerate(row):
                    ws.cell(row=row0 + i, column=1 + j).value = v
            for i, row in enumerate(tb_z):
                for j, v in enumerate(row):
                    ws.cell(row=row0 + i, column=11 + j).value = v
            logger.info("export_report: data written for '{}'", title)

            maxr = max(len(tb_b), len(tb_z))
            endr = row0 + maxr - 1

            for off in range(maxr):
                r = row0 + off
                if r < 8: continue
                for cb, cz in ((3, 12), (4, 13), (5, 14)):
                    if ws.cell(r, cb).value != ws.cell(r, cz).value:
                        ws.cell(r, cb).font = red
                        ws.cell(r, cz).font = red
            logger.info("export_report: mismatches highlighted for '{}'", title)

            for col in ("P", "Q", "R"):
                rng = f"{col}8:{col}{endr}"
                rule = CellIsRule(operator='notEqual', formula=['0'], font=red)
                ws.conditional_formatting.add(rng, rule)
            logger.info("export_report: conditional formatting for '{}'", title)

        except Exception:
            logger.error("export_report: ошибка на листе '{}', пропускаем'", title)
            continue

    try:
        wb.remove(wb[base])
        wb.save(OUTPUT)
        logger.info("export_report: saved output '{}'", OUTPUT)
    except Exception:
        logger.critical("export_report: не удалось сохранить '{}', выходим'", OUTPUT)
        raise
    finally:
        if wb:
            wb.close()
            logger.info("export_report: workbook closed")

    logger.info("export_report: exit with '{}'", OUTPUT)
    return OUTPUT
