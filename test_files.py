import os
from openpyxl import Workbook

EXAMPLE_DIR = "example_xlsx"

MONTH_MAPPING = {
    1: "январь", 2: "февраль", 3: "март",
    4: "апрель", 5: "май", 6: "июнь",
    7: "июль", 8: "август", 9: "сентябрь",
    10: "октябрь", 11: "ноябрь", 12: "декабрь"
}

CONTENT_KEYWORDS = {
    "zup": "Ведомость в кассу",
    "buh": "Выдача наличных"
}

FILENAME_KEYWORDS = {
    "zup": "зуп",
    "buh": "бух"
}

scenarios = []

MONTH_FOR_TEST = 1

for category in ("zup", "buh"):
    for class_mode in ("filename", "content"):
        for month_mode in (
                "filename_and_d4_date",
                "filename_and_content_date",
                "content_d4",
                "content_content"
        ):
            for day in (10, 25):
                scenarios.append({
                    "category": category,
                    "class_mode": class_mode,
                    "month_mode": month_mode,
                    "month": MONTH_FOR_TEST,
                    "day": day
                })
        for month_mode in ("filename_no_date", "no_month_no_date"):
            scenarios.append({
                "category": category,
                "class_mode": class_mode,
                "month_mode": month_mode,
                "month": MONTH_FOR_TEST,
                "day": None
            })

scenarios.append({
    "category": None,
    "class_mode": "none",
    "month_mode": "no_month_no_date",
    "month": None,
    "day": None
})


def create_example_files():
    os.makedirs(EXAMPLE_DIR, exist_ok=True)

    for sc in scenarios:
        category = sc["category"]
        class_mode = sc["class_mode"]
        month_mode = sc["month_mode"]
        month = sc["month"]
        day = sc["day"]

        parts = []
        if class_mode == "filename" and category:
            parts.append(FILENAME_KEYWORDS[category])
        if month_mode in ("filename_and_d4_date", "filename_and_content_date", "filename_no_date") and month:
            parts.append(MONTH_MAPPING[month])
        parts.append(class_mode)
        parts.append(month_mode)
        if day is not None:
            parts.append(f"day{day}")

        filename = "_".join(parts) + ".xlsx"
        path = os.path.join(EXAMPLE_DIR, filename)

        wb = Workbook()
        ws = wb.active

        if class_mode == "content" and category:
            ws["E4"] = CONTENT_KEYWORDS[category]

        if month_mode == "filename_and_d4_date":
            ws["D4"] = f"{day:02d}.{month:02d}.2025"
        elif month_mode == "filename_and_content_date":
            ws["A1"] = f"{day:02d}.{month:02d}.2025"
        elif month_mode == "content_d4":
            ws["D4"] = f"{day:02d}.{month:02d}.2025"
        elif month_mode == "content_content":
            ws["B2"] = f"{day:02d}.{month:02d}.2025"

        wb.save(path)

    print(f"Сгенерировано {len(scenarios)} примеров в папке ./{EXAMPLE_DIR}/")


if __name__ == "__main__":
    create_example_files()
